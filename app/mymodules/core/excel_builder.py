import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import IconSetRule
import os

def apply_borders(ws, topleft: str, len_cols: int, len_rows: int):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    top_row = int(topleft[1:])  # Extract row number
    left_col = openpyxl.utils.column_index_from_string(topleft[0])  # Convert column letter to index

    for row in range(top_row, top_row + len_rows):
        for col in range(left_col, left_col + len_cols):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border  # Add border

    for col in range(left_col, left_col + len_cols):  # Adjusting width for columns A to G (6 columns plus one for the main header)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

def apply_avg_format(ws, topleft: str, column_count: int):
    col_letter = topleft[0]
    row_number = int(topleft[1:])
    col_index = openpyxl.utils.column_index_from_string(col_letter)

    fill_color = PatternFill(start_color='cbd7f0', end_color='cbd7f0', fill_type='solid')

    for i in range(column_count):
        cell = ws.cell(row=row_number, column=col_index + i)
        cell.font = Font(bold=True)
        if i == 0:
            cell.fill = fill_color

def build_header(ws, diff_cell_content: str, topleft: str):
    top_row = int(topleft[1:])  # Extract row number
    left_col = openpyxl.utils.column_index_from_string(topleft[0])  # Convert column letter to index

    ws.merge_cells(start_row=top_row, start_column=left_col, end_row=top_row + 1, end_column=left_col)
    header_cell = ws.cell(row=top_row, column=left_col)
    header_cell.value = diff_cell_content

    ws.merge_cells(start_row=top_row, start_column=left_col + 1, end_row=top_row, end_column=left_col + 6)
    indicators_cell = ws.cell(row=top_row, column=left_col + 1)
    indicators_cell.value = 'INDICADORES'

    indicators = ['META SALDO', 'CREC. SALDO', 'META CLIENTES', 'CREC. CLIENTES', 'META DESEMBOLSOS', 'DESEMBOLSOS']
    
    for i, indicator in enumerate(indicators):
        ws.cell(row=top_row + 1, column=left_col + 1 + i).value = indicator

    fill_color = "cbd7f0"
    font_style = Font(size=11, bold=True)  # Changed font size to 11
    
    header_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_cell.font = font_style
    header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable text wrapping
    
    indicators_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    indicators_cell.font = font_style
    indicators_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable text wrapping

    for i in range(len(indicators)):
        indicator_cell = ws.cell(row=top_row + 1, column=left_col + 1 + i)
        indicator_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        indicator_cell.font = font_style
        indicator_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # Enable text wrapping

    apply_borders(ws, topleft, len_cols=7, len_rows=2)  # Main header and indicators section

def build_data_row(ws, data: list, topleft: str, traffic_light_columns: list=[4]):

    top_row = int(topleft[1:])
    left_col = openpyxl.utils.column_index_from_string(topleft[0])
    
    for i in range(len(data)):
        col_index = left_col + i
        
        ws.cell(row=top_row, column=col_index).value = data[i]
        
        ws.cell(row=top_row, column=col_index).alignment = Alignment(horizontal='center', vertical='center')

    for i in traffic_light_columns:
        icon_set_rule = IconSetRule('3TrafficLights1', type='num', values=[-10000000, 10000000], showValue=None)
        ws.conditional_formatting.add(f"{openpyxl.utils.get_column_letter(left_col + i)}{top_row}:{openpyxl.utils.get_column_letter(left_col + i)}{top_row}", icon_set_rule)

    apply_borders(ws, f"{topleft[0]}{top_row}", len_cols=len(data), len_rows=1)
    
def insert_timeline(ws, timeline, col_start: str, row_start: int, traffic_light_cols = [4]):
    idx_row = row_start
    for row in timeline:
        build_data_row(ws, list(row.values()), f'{col_start.upper()}{idx_row}', traffic_light_cols)
        idx_row += 1

    return idx_row

def for_target_with_months(wb: openpyxl.Workbook, data):
    ws = wb.create_sheet(title='Cuadro1')
    build_header(ws, data['username'], 'A1')
    build_data_row(ws, list(data['timeline'][-1].values()), 'A3', [4])

    ws2 = wb.create_sheet(title='Cuadro4')
    build_header(ws2, data['username'], 'A1')
    insert_timeline(ws2, data['timeline'], 'A', 3)

def for_all_months_pairs_avg(wb: openpyxl.Workbook, data):
    ws = wb.create_sheet(title='Cuadro3')
    build_header(ws, 'PROMEDIO DE SUS PARES', 'A1')
    build_data_row(ws, list(data['timeline'][-1].values()), 'A3', [4])

    ws = wb.create_sheet(title='Cuadro7')
    build_header(ws, 'PROMEDIO DE SUS PARES', 'A1')
    insert_timeline(ws, data['timeline'], 'A', 3)

def for_full_avg_target(wb: openpyxl.Workbook, data):
    ws = wb.create_sheet(title='Cuadro5&8')
    build_header(ws, data['username'], 'A1')
    values = list(data['average'][-1].values())
    values[0] = 'PROMEDIO'
    build_data_row(ws, values, 'A3', [4])
    apply_avg_format(ws, 'A3', 7)

def for_full_avg_pairs(wb: openpyxl.Workbook, data):
    ws = wb.create_sheet(title='Cuadro9')
    build_header(ws, 'PROMEDIO DE SUS PARES', 'A1')
    values = list(data['average'][-1].values())
    values[0] = 'PROMEDIO'
    build_data_row(ws, values, 'A3', [4])
    apply_avg_format(ws, 'A3', 7)

def for_all_pairsxmonths(wb: openpyxl.Workbook, data):
    ws = wb.create_sheet(title='Cuadros6')

    idx_row = 1
    for pair in data:
        build_header(ws, pair['username'], f'A{idx_row}')
        idx_row += 2
        idx_row = insert_timeline(ws, pair['timeline'], 'A', idx_row)
        avg = list(pair['average'][-1].values())
        avg[0] = 'PROMEDIO'
        build_data_row(ws, avg, f'A{idx_row}')
        apply_avg_format(ws, f'A{idx_row}', 7)
        idx_row += 2

def build_tables(tables, download_path: str):
    table_factories = {
        'all_months_target': for_target_with_months,
        'all_months_pairs_avg': for_all_months_pairs_avg,
        'full_avg_target': for_full_avg_target,
        'full_avg_pairs': for_full_avg_pairs,
        'pairs_all_months': for_all_pairsxmonths,
    }


    wb = openpyxl.Workbook()
    sheet0 = wb.sheetnames[0]
    for key, value in tables['tables'].items():
        table_factories[key](wb, value)

    wb.remove(wb[sheet0])

    base_path, extension = os.path.splitext(download_path)
    counter = 1

    while os.path.exists(download_path):
        download_path = f"{base_path} ({counter}){extension}"
        counter += 1

    wb.save(download_path)

    return download_path